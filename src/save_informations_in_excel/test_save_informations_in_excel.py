import os
from openpyxl import load_workbook
import pytest
from faker import Faker
from save_informations_in_excel import ExcelRobot

fake = Faker()


@pytest.fixture
def excel_open_and_delete():
    ExcelRobot()
    yield
    os.remove(os.path.abspath("data/air-travels.xlsx"))


def make_mock_extracted_informations():
    mock_extracted_informations = []
    for _ in list(range(5)):
        mock_extracted_informations.append(
            {
                "price": fake.pricetag(),
                "departure time": fake.date(),
                "disembarkation time": fake.date(),
            },
        )
    return mock_extracted_informations


@pytest.mark.excel_process
def test_insert_extracted_informations_in_excel_file(excel_open_and_delete):
    mock_extracted_informations = make_mock_extracted_informations()

    excel_robot = ExcelRobot()
    excel_robot.insert_informations_in_excel(mock_extracted_informations)
    excel_robot.save_excel()

    excel_file = load_workbook(os.path.abspath("data/air-travels.xlsx"))
    sheet = excel_file.active

    index = 0
    for row in sheet.iter_rows(min_row=2, min_col=1, max_row=6, max_col=3):
        assert row[0].value == mock_extracted_informations[index]["price"]
        assert row[1].value == mock_extracted_informations[index]["departure time"]
        assert row[2].value == mock_extracted_informations[index]["disembarkation time"]
        index += 1


@pytest.mark.excel_process
def test_create_blank_excel_file(excel_open_and_delete):
    mock_excel_file_path = os.path.abspath("data/air-travels.xlsx")
    assert os.path.isfile(mock_excel_file_path) == True


@pytest.mark.excel_process
def test_not_create_new_excel_file_if_exists_one():
    ExcelRobot()
    mock_excel_file_path = os.path.abspath("data/air-travels.xlsx")
    excel_date_creation = os.stat(mock_excel_file_path)
    ExcelRobot()
    assert excel_date_creation.st_mtime == os.stat(mock_excel_file_path).st_mtime
    os.remove(os.path.abspath("data/air-travels.xlsx"))
