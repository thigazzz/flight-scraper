import os
from typing import Dict
from openpyxl import load_workbook, Workbook


class ExcelRobot:
    excel_file_path = os.path.abspath("data/air-travels.xlsx")

    def __init__(self) -> None:
        self.create_blank_excel_file()
        self.excel = self.get_excel()

    def create_blank_excel_file(self):
        if self.is_file_excel_exists() == False:
            Workbook().save(self.excel_file_path)

    def is_file_excel_exists(self) -> bool:
        return os.path.isfile(self.excel_file_path)

    def get_excel(self):
        return load_workbook(self.excel_file_path)

    def get_actual_sheet(self):
        return self.excel.active

    def insert_informations_in_excel(self, informations: Dict):
        index = 0
        for row in self.iterate_rows_of_length_informations(informations):
            self.insert_in_cell(row[0], informations[index]["price"])
            self.insert_in_cell(row[1], informations[index]["departure time"])
            self.insert_in_cell(row[2], informations[index]["disembarkation time"])
            index += 1

    def iterate_rows_of_length_informations(self, data):
        sheet = self.get_actual_sheet()
        return sheet.iter_rows(
            min_row=2, min_col=1, max_row=len(data) + 1, max_col=len(data[0])
        )

    def insert_in_cell(self, row, data):
        row.value = data

    def save_excel(self):
        self.excel.save(self.excel_file_path)
